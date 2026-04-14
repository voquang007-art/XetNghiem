# -*- coding: utf-8 -*-
from __future__ import annotations

import os
import uuid
from datetime import date, datetime, timedelta
from io import BytesIO
from typing import Any

from fastapi import APIRouter, Depends, Form, HTTPException, Request, status
from sqlalchemy import text
from sqlalchemy.orm import Session
from starlette.responses import RedirectResponse, StreamingResponse
from starlette.templating import Jinja2Templates

from ..config import settings
from ..database import get_db
from ..models import Roles, RoleCode, Units, UserRoles, UserUnitMemberships, Users
from ..security.deps import login_required

router = APIRouter(prefix="/leave-schedule", tags=["leave_schedule"])
templates = Jinja2Templates(directory=os.path.join(os.path.dirname(__file__), "..", "templates"))

ROLE_TRUONG_KHOA = RoleCode.ROLE_TRUONG_KHOA.value
ROLE_PHO_TRUONG_KHOA = RoleCode.ROLE_PHO_TRUONG_KHOA.value
ROLE_KTV_TRUONG = RoleCode.ROLE_KY_THUAT_VIEN_TRUONG.value
ROLE_ADMIN = RoleCode.ROLE_ADMIN.value
ROLE_QL_CONG_VIEC = RoleCode.ROLE_QL_CONG_VIEC.value

FUNCTIONAL_ROLE_CODES = {
    RoleCode.ROLE_QL_CHAT_LUONG.value,
    RoleCode.ROLE_QL_KY_THUAT.value,
    RoleCode.ROLE_QL_AN_TOAN.value,
    RoleCode.ROLE_QL_VAT_TU.value,
    RoleCode.ROLE_QL_TRANG_THIET_BI.value,
    RoleCode.ROLE_QL_MOI_TRUONG.value,
    RoleCode.ROLE_QL_CNTT.value,
}

MANAGER_ROLE_CODES = {
    ROLE_ADMIN,
    ROLE_TRUONG_KHOA,
    ROLE_PHO_TRUONG_KHOA,
    ROLE_KTV_TRUONG,
}

APPROVER_ROLE_CODES = {
    ROLE_TRUONG_KHOA,
    ROLE_PHO_TRUONG_KHOA,
}

LEAVE_TYPE_META = {
    "F": {"label": "Nghỉ phép năm", "max_year": 12, "max_month": None},
    "P": {"label": "Nghỉ phép tháng", "max_year": None, "max_month": 4},
    "NL": {"label": "Nghỉ lễ, tết", "max_year": 12, "max_month": None},
    "H": {"label": "Đi học", "max_year": None, "max_month": None},
    "CT": {"label": "Đi công tác", "max_year": None, "max_month": None},
}


CREATE_LEAVE_REQUESTS_SQL = """
CREATE TABLE IF NOT EXISTS leave_requests (
    id TEXT PRIMARY KEY,
    user_id TEXT NOT NULL,
    user_name TEXT,
    unit_id TEXT,
    unit_name TEXT,
    group_key TEXT NOT NULL,
    group_label TEXT NOT NULL,
    leave_type TEXT NOT NULL,
    symbol TEXT NOT NULL,
    start_date TEXT NOT NULL,
    end_date TEXT NOT NULL,
    day_count REAL NOT NULL DEFAULT 0,
    reason TEXT,
    status TEXT NOT NULL DEFAULT 'PENDING',
    approved_by_id TEXT,
    approved_by_name TEXT,
    approved_role TEXT,
    approved_at TEXT,
    created_at TEXT NOT NULL,
    updated_at TEXT NOT NULL
)
"""

CREATE_LEAVE_ADJUSTMENTS_SQL = """
CREATE TABLE IF NOT EXISTS leave_year_adjustments (
    id TEXT PRIMARY KEY,
    user_id TEXT NOT NULL,
    leave_year INTEGER NOT NULL,
    extra_days INTEGER NOT NULL DEFAULT 0,
    reason TEXT,
    created_by_id TEXT,
    created_at TEXT NOT NULL
)
"""

CREATE_INDEXES_SQL = [
    "CREATE INDEX IF NOT EXISTS idx_leave_requests_user_status ON leave_requests (user_id, status)",
    "CREATE INDEX IF NOT EXISTS idx_leave_requests_group_dates ON leave_requests (group_key, start_date, end_date)",
    "CREATE INDEX IF NOT EXISTS idx_leave_adjustments_user_year ON leave_year_adjustments (user_id, leave_year)",
]


def _now_vn() -> datetime:
    return datetime.utcnow() + timedelta(hours=7)


def _today_vn() -> date:
    return _now_vn().date()


def _ensure_tables(db: Session) -> None:
    db.execute(text(CREATE_LEAVE_REQUESTS_SQL))
    db.execute(text(CREATE_LEAVE_ADJUSTMENTS_SQL))
    for sql in CREATE_INDEXES_SQL:
        db.execute(text(sql))
    db.commit()


def _parse_date(value: str, field_name: str) -> date:
    try:
        return datetime.strptime((value or "").strip(), "%Y-%m-%d").date()
    except Exception as exc:
        raise HTTPException(status_code=400, detail=f"{field_name} không hợp lệ.") from exc


def _date_to_str(value: date | None) -> str | None:
    return value.isoformat() if value else None


def _dt_to_str(value: datetime | None) -> str | None:
    return value.strftime("%Y-%m-%d %H:%M:%S") if value else None


def _month_cutoff(any_day: date) -> int:
    if any_day.month == 2:
        return 23
    if any_day.month in {4, 6, 9, 11}:
        return 25
    return 26


def _iter_workdays(start_date: date, end_date: date):
    current = start_date
    while current <= end_date:
        if current.weekday() != 6:  # Chủ nhật nghỉ
            yield current
        current += timedelta(days=1)


def _count_workdays(start_date: date, end_date: date) -> int:
    return sum(1 for _ in _iter_workdays(start_date, end_date))


def _row_to_dict(row: Any) -> dict[str, Any]:
    if hasattr(row, "_mapping"):
        return dict(row._mapping)
    return dict(row)


def _get_role_codes(db: Session, user_id: str) -> set[str]:
    rows = (
        db.query(Roles.code)
        .join(UserRoles, UserRoles.role_id == Roles.id)
        .filter(UserRoles.user_id == user_id)
        .all()
    )
    return {str(getattr(code, "value", code)) for (code,) in rows}


def _primary_unit(db: Session, user_id: str) -> Units | None:
    membership = (
        db.query(UserUnitMemberships)
        .join(Units, Units.id == UserUnitMemberships.unit_id)
        .filter(
            UserUnitMemberships.user_id == user_id,
            UserUnitMemberships.is_active.is_(True),
            UserUnitMemberships.is_primary.is_(True),
        )
        .order_by(UserUnitMemberships.id.desc())
        .first()
    )
    return membership.unit if membership and membership.unit else None


def _resolve_group_bucket(db: Session, user: Users) -> tuple[str, str, Units | None]:
    role_codes = _get_role_codes(db, user.id)
    unit = _primary_unit(db, user.id)

    if role_codes & {ROLE_TRUONG_KHOA, ROLE_PHO_TRUONG_KHOA, ROLE_KTV_TRUONG}:
        return ("LAB_LEADERSHIP", "Nhóm Trưởng khoa / Phó khoa / KTV trưởng", unit)

    if ROLE_QL_CONG_VIEC in role_codes:
        return ("QL_CONG_VIEC", "Nhóm Quản lý công việc", unit)

    if role_codes & FUNCTIONAL_ROLE_CODES:
        return ("QL_CHUC_NANG", "Nhóm Quản lý chức năng", unit)

    if unit is not None:
        return (f"UNIT::{unit.id}", unit.ten_don_vi or "Nhóm", unit)

    return (f"USER::{user.id}", user.full_name or user.username or "Người dùng", unit)


def _can_admin_extra_days(db: Session, user: Users) -> bool:
    return ROLE_ADMIN in _get_role_codes(db, user.id)


def _can_adjust_after_cutoff(db: Session, user: Users) -> bool:
    return bool(_get_role_codes(db, user.id) & MANAGER_ROLE_CODES)


def _can_approve(db: Session, user: Users) -> bool:
    return bool(_get_role_codes(db, user.id) & APPROVER_ROLE_CODES)


def _role_label_for_user(db: Session, user: Users) -> str:
    role_codes = _get_role_codes(db, user.id)
    if ROLE_TRUONG_KHOA in role_codes:
        return "Trưởng khoa"
    if ROLE_PHO_TRUONG_KHOA in role_codes:
        return "Phó khoa"
    if ROLE_KTV_TRUONG in role_codes:
        return "Kỹ thuật viên trưởng"
    if ROLE_ADMIN in role_codes:
        return "Admin"
    return "Người dùng"


def _is_cutoff_locked(db: Session, user: Users, target_month: date | None = None) -> bool:
    current_day = _today_vn()
    month_base = target_month or current_day
    cutoff = _month_cutoff(month_base)
    if current_day.day <= cutoff:
        return False
    return not _can_adjust_after_cutoff(db, user)


def _sum_extra_days(db: Session, user_id: str, leave_year: int) -> int:
    row = db.execute(
        text(
            """
            SELECT COALESCE(SUM(extra_days), 0) AS total_extra
            FROM leave_year_adjustments
            WHERE user_id = :user_id AND leave_year = :leave_year
            """
        ),
        {"user_id": user_id, "leave_year": leave_year},
    ).fetchone()
    if not row:
        return 0
    return int(_row_to_dict(row).get("total_extra") or 0)


def _sum_days_by_year(db: Session, user_id: str, leave_type: str, leave_year: int, statuses: tuple[str, ...]) -> float:
    row = db.execute(
        text(
            f"""
            SELECT COALESCE(SUM(day_count), 0) AS total_days
            FROM leave_requests
            WHERE user_id = :user_id
              AND leave_type = :leave_type
              AND status IN ({','.join([f':st{i}' for i in range(len(statuses))])})
              AND substr(start_date, 1, 4) = :leave_year
            """
        ),
        {
            "user_id": user_id,
            "leave_type": leave_type,
            "leave_year": str(leave_year),
            **{f"st{i}": statuses[i] for i in range(len(statuses))},
        },
    ).fetchone()
    if not row:
        return 0.0
    return float(_row_to_dict(row).get("total_days") or 0)


def _sum_days_by_month(db: Session, user_id: str, leave_type: str, year_month: str, statuses: tuple[str, ...]) -> float:
    row = db.execute(
        text(
            f"""
            SELECT COALESCE(SUM(day_count), 0) AS total_days
            FROM leave_requests
            WHERE user_id = :user_id
              AND leave_type = :leave_type
              AND status IN ({','.join([f':st{i}' for i in range(len(statuses))])})
              AND substr(start_date, 1, 7) = :year_month
            """
        ),
        {
            "user_id": user_id,
            "leave_type": leave_type,
            "year_month": year_month,
            **{f"st{i}": statuses[i] for i in range(len(statuses))},
        },
    ).fetchone()
    if not row:
        return 0.0
    return float(_row_to_dict(row).get("total_days") or 0)


def _validate_quota(db: Session, user: Users, leave_type: str, start_date: date, end_date: date, requested_days: int) -> None:
    meta = LEAVE_TYPE_META[leave_type]
    statuses = ("PENDING", "APPROVED")

    if leave_type == "F":
        leave_year = start_date.year
        total_extra = _sum_extra_days(db, user.id, leave_year)
        used_days = _sum_days_by_year(db, user.id, leave_type, leave_year, statuses)
        max_days = int(meta["max_year"] or 0) + total_extra
        if used_days + requested_days > max_days:
            raise HTTPException(
                status_code=400,
                detail=(
                    f"Nghỉ phép năm vượt giới hạn. Đã dùng/đang chờ: {used_days:g} ngày; "
                    f"được phép: {max_days:g} ngày."
                ),
            )

    if leave_type == "NL":
        leave_year = start_date.year
        used_days = _sum_days_by_year(db, user.id, leave_type, leave_year, statuses)
        max_days = int(meta["max_year"] or 0)
        if used_days + requested_days > max_days:
            raise HTTPException(
                status_code=400,
                detail=(
                    f"Nghỉ lễ, tết vượt mức cấu hình {max_days:g} ngày/năm."
                ),
            )

    if leave_type == "P":
        month_map: dict[str, int] = {}
        for d in _iter_workdays(start_date, end_date):
            ym = d.strftime("%Y-%m")
            month_map[ym] = month_map.get(ym, 0) + 1
        for ym, add_days in month_map.items():
            used_days = _sum_days_by_month(db, user.id, leave_type, ym, statuses)
            max_days = int(meta["max_month"] or 0)
            if used_days + add_days > max_days:
                raise HTTPException(
                    status_code=400,
                    detail=(
                        f"Nghỉ phép tháng vượt giới hạn 4 ngày công/tháng. "
                        f"Tháng {ym} đã dùng/đang chờ: {used_days:g} ngày."
                    ),
                )


def _find_group_conflict(
    db: Session,
    user: Users,
    group_key: str,
    leave_type: str,
    start_date: date,
    end_date: date,
) -> dict[str, Any] | None:
    rows = db.execute(
        text(
            """
            SELECT *
            FROM leave_requests
            WHERE group_key = :group_key
              AND status IN ('PENDING', 'APPROVED')
              AND NOT (end_date < :start_date OR start_date > :end_date)
            ORDER BY start_date ASC, created_at ASC
            """
        ),
        {
            "group_key": group_key,
            "start_date": _date_to_str(start_date),
            "end_date": _date_to_str(end_date),
        },
    ).fetchall()

    wanted_dates = {d.isoformat() for d in _iter_workdays(start_date, end_date)}
    for row in rows:
        item = _row_to_dict(row)
        if item.get("user_id") == user.id:
            return {
                "message": "Khoảng thời gian này đã có đăng ký nghỉ trước đó của chính người dùng.",
                "row": item,
            }
        existing_dates = {
            d.isoformat()
            for d in _iter_workdays(_parse_date(item["start_date"], "Ngày bắt đầu"), _parse_date(item["end_date"], "Ngày kết thúc"))
        }
        if wanted_dates & existing_dates:
            return {
                "message": (
                    "Ngày đăng ký bị khóa do trong cùng nhóm đã có người đăng ký nghỉ hoặc nghỉ loại khóa cứng."
                    if leave_type == "P"
                    else "Nhóm này đã có người đăng ký nghỉ trong cùng ngày nên không thể đăng ký thêm."
                ),
                "row": item,
            }
    return None


def _format_request_row(item: dict[str, Any]) -> dict[str, Any]:
    meta = LEAVE_TYPE_META.get(item.get("leave_type") or "", {})
    status = item.get("status") or ""
    status_label = {
        "PENDING": "Chờ duyệt",
        "APPROVED": "Đã duyệt",
        "CANCELLED": "Đã hủy",
        "REJECTED": "Từ chối",
    }.get(status, status)
    item = dict(item)
    item["leave_type_label"] = meta.get("label") or item.get("leave_type")
    item["status_label"] = status_label
    item["can_cancel"] = status == "PENDING"
    return item


def _load_requests(db: Session, user: Users, viewer_is_manager: bool, tab: str) -> list[dict[str, Any]]:
    if viewer_is_manager:
        sql = "SELECT * FROM leave_requests ORDER BY start_date DESC, created_at DESC"
        rows = db.execute(text(sql)).fetchall()
    else:
        rows = db.execute(
            text(
                "SELECT * FROM leave_requests WHERE user_id = :user_id ORDER BY start_date DESC, created_at DESC"
            ),
            {"user_id": user.id},
        ).fetchall()
    items = [_format_request_row(_row_to_dict(r)) for r in rows]
    if tab == "pending":
        return [x for x in items if x.get("status") == "PENDING"]
    if tab == "approved":
        return [x for x in items if x.get("status") == "APPROVED"]
    return items


def _build_balance_cards(db: Session, user: Users) -> list[dict[str, Any]]:
    current_year = _today_vn().year
    current_month = _today_vn().strftime("%Y-%m")
    extra_f = _sum_extra_days(db, user.id, current_year)
    used_f = _sum_days_by_year(db, user.id, "F", current_year, ("PENDING", "APPROVED"))
    used_p = _sum_days_by_month(db, user.id, "P", current_month, ("PENDING", "APPROVED"))
    used_nl = _sum_days_by_year(db, user.id, "NL", current_year, ("PENDING", "APPROVED"))
    return [
        {
            "code": "F",
            "label": "Nghỉ phép năm",
            "used": used_f,
            "max": 12 + extra_f,
            "note": f"Gồm 12 ngày chuẩn + {extra_f} ngày admin cộng thêm" if extra_f else "Mặc định 12 ngày làm việc/năm",
        },
        {
            "code": "P",
            "label": "Nghỉ phép tháng",
            "used": used_p,
            "max": 4,
            "note": f"Áp dụng cho tháng {current_month}",
        },
        {
            "code": "NL",
            "label": "Nghỉ lễ, tết",
            "used": used_nl,
            "max": 12,
            "note": "Theo mức cấu hình hiện tại của khoa",
        },
        {
            "code": "H",
            "label": "Đi học",
            "used": 0,
            "max": None,
            "note": "Không giới hạn số ngày trong module này",
        },
        {
            "code": "CT",
            "label": "Đi công tác",
            "used": 0,
            "max": None,
            "note": "Không giới hạn số ngày trong module này",
        },
    ]


@router.get("")
def leave_schedule_index(
    request: Request,
    tab: str = "leave",
    list_mode: str = "all",
    month: str = "",
    db: Session = Depends(get_db),
):
    _ensure_tables(db)
    user = login_required(request, db)
    current_tab = "roster" if tab == "roster" else "leave"
    viewer_is_manager = _can_adjust_after_cutoff(db, user)
    approver = _can_approve(db, user)
    balances = _build_balance_cards(db, user)
    requests = _load_requests(db, user, viewer_is_manager or approver or _can_admin_extra_days(db, user), list_mode)
    can_submit = not _is_cutoff_locked(db, user)
    group_key, group_label, primary_unit = _resolve_group_bucket(db, user)

    msg = request.query_params.get("msg", "")
    error = request.query_params.get("error", "")
    current_month = month or _today_vn().strftime("%Y-%m")

    return templates.TemplateResponse(
        "leave_schedule.html",
        {
            "request": request,
            "app_name": settings.APP_NAME,
            "company_name": settings.COMPANY_NAME,
            "current_tab": current_tab,
            "current_list_mode": list_mode,
            "requests": requests,
            "balances": balances,
            "can_submit": can_submit,
            "can_approve": approver,
            "can_admin_adjust": _can_admin_extra_days(db, user),
            "is_manager": viewer_is_manager,
            "group_key": group_key,
            "group_label": group_label,
            "primary_unit_name": getattr(primary_unit, "ten_don_vi", "") if primary_unit else "",
            "leave_type_meta": LEAVE_TYPE_META,
            "today_str": _today_vn().isoformat(),
            "current_month": current_month,
            "cutoff_day": _month_cutoff(_today_vn()),
            "message": msg,
            "error": error,
        },
    )


@router.post("/add")
def create_leave_request(
    request: Request,
    leave_type: str = Form(...),
    start_date: str = Form(...),
    end_date: str = Form(...),
    reason: str = Form(""),
    db: Session = Depends(get_db),
):
    _ensure_tables(db)
    user = login_required(request, db)

    if _is_cutoff_locked(db, user):
        return RedirectResponse(
            url="/leave-schedule?error=Đã qua thời hạn đăng ký trong tháng. Chỉ Trưởng khoa, Phó khoa, Kỹ thuật viên trưởng hoặc Admin mới được điều chỉnh.",
            status_code=302,
        )

    leave_type = (leave_type or "").strip().upper()
    if leave_type not in LEAVE_TYPE_META:
        raise HTTPException(status_code=400, detail="Loại nghỉ không hợp lệ.")

    start_dt = _parse_date(start_date, "Ngày bắt đầu")
    end_dt = _parse_date(end_date, "Ngày kết thúc")
    if end_dt < start_dt:
        raise HTTPException(status_code=400, detail="Ngày kết thúc không được nhỏ hơn ngày bắt đầu.")

    requested_days = _count_workdays(start_dt, end_dt)
    if requested_days <= 0:
        raise HTTPException(status_code=400, detail="Khoảng thời gian đăng ký không có ngày công hợp lệ.")

    group_key, group_label, unit = _resolve_group_bucket(db, user)
    _validate_quota(db, user, leave_type, start_dt, end_dt, requested_days)

    conflict = _find_group_conflict(db, user, group_key, leave_type, start_dt, end_dt)
    if conflict:
        row = conflict.get("row") or {}
        detail = conflict["message"]
        if row:
            detail += f" Người đã đăng ký: {row.get('user_name') or '-'}; thời gian: {row.get('start_date')} đến {row.get('end_date')}"
        return RedirectResponse(url=f"/leave-schedule?error={detail}", status_code=302)

    now_str = _dt_to_str(_now_vn())
    db.execute(
        text(
            """
            INSERT INTO leave_requests (
                id, user_id, user_name, unit_id, unit_name, group_key, group_label,
                leave_type, symbol, start_date, end_date, day_count, reason,
                status, created_at, updated_at
            ) VALUES (
                :id, :user_id, :user_name, :unit_id, :unit_name, :group_key, :group_label,
                :leave_type, :symbol, :start_date, :end_date, :day_count, :reason,
                'PENDING', :created_at, :updated_at
            )
            """
        ),
        {
            "id": str(uuid.uuid4()),
            "user_id": user.id,
            "user_name": user.full_name or user.username,
            "unit_id": getattr(unit, "id", None),
            "unit_name": getattr(unit, "ten_don_vi", None),
            "group_key": group_key,
            "group_label": group_label,
            "leave_type": leave_type,
            "symbol": leave_type,
            "start_date": _date_to_str(start_dt),
            "end_date": _date_to_str(end_dt),
            "day_count": requested_days,
            "reason": (reason or "").strip(),
            "created_at": now_str,
            "updated_at": now_str,
        },
    )
    db.commit()
    return RedirectResponse(url="/leave-schedule?msg=Đã ghi nhận đăng ký nghỉ.", status_code=302)


@router.post("/approve")
def approve_leave_request(
    request: Request,
    request_id: str = Form(...),
    db: Session = Depends(get_db),
):
    _ensure_tables(db)
    user = login_required(request, db)
    if not _can_approve(db, user):
        raise HTTPException(status_code=403, detail="Chỉ Trưởng khoa hoặc Phó khoa được phê duyệt.")

    row = db.execute(text("SELECT * FROM leave_requests WHERE id = :id"), {"id": request_id}).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Không tìm thấy đăng ký nghỉ.")

    item = _row_to_dict(row)
    if item.get("status") != "PENDING":
        return RedirectResponse(url="/leave-schedule?error=Phiếu này đã được xử lý.", status_code=302)

    now_str = _dt_to_str(_now_vn())
    db.execute(
        text(
            """
            UPDATE leave_requests
            SET status = 'APPROVED',
                approved_by_id = :approved_by_id,
                approved_by_name = :approved_by_name,
                approved_role = :approved_role,
                approved_at = :approved_at,
                updated_at = :updated_at
            WHERE id = :id AND status = 'PENDING'
            """
        ),
        {
            "id": request_id,
            "approved_by_id": user.id,
            "approved_by_name": user.full_name or user.username,
            "approved_role": _role_label_for_user(db, user),
            "approved_at": now_str,
            "updated_at": now_str,
        },
    )
    db.commit()
    return RedirectResponse(url="/leave-schedule?list_mode=pending&msg=Đã phê duyệt đăng ký nghỉ.", status_code=302)


@router.post("/cancel")
def cancel_leave_request(
    request: Request,
    request_id: str = Form(...),
    db: Session = Depends(get_db),
):
    _ensure_tables(db)
    user = login_required(request, db)
    row = db.execute(text("SELECT * FROM leave_requests WHERE id = :id"), {"id": request_id}).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Không tìm thấy đăng ký nghỉ.")
    item = _row_to_dict(row)
    if item.get("status") != "PENDING":
        return RedirectResponse(url="/leave-schedule?error=Chỉ phiếu đang chờ duyệt mới được hủy.", status_code=302)

    can_cancel = item.get("user_id") == user.id or _can_adjust_after_cutoff(db, user) or _can_admin_extra_days(db, user)
    if not can_cancel:
        raise HTTPException(status_code=403, detail="Bạn không có quyền hủy phiếu này.")

    db.execute(
        text(
            "UPDATE leave_requests SET status = 'CANCELLED', updated_at = :updated_at WHERE id = :id"
        ),
        {"id": request_id, "updated_at": _dt_to_str(_now_vn())},
    )
    db.commit()
    return RedirectResponse(url="/leave-schedule?msg=Đã hủy đăng ký nghỉ.", status_code=302)


@router.post("/adjust-annual")
def adjust_annual_leave(
    request: Request,
    user_id: str = Form(...),
    leave_year: int = Form(...),
    extra_days: int = Form(...),
    reason: str = Form(""),
    db: Session = Depends(get_db),
):
    _ensure_tables(db)
    user = login_required(request, db)
    if not _can_admin_extra_days(db, user):
        raise HTTPException(status_code=403, detail="Chỉ Admin mới được cộng thêm ngày phép năm.")
    target_user = db.get(Users, user_id)
    if not target_user:
        raise HTTPException(status_code=404, detail="Không tìm thấy người dùng.")
    if extra_days <= 0:
        raise HTTPException(status_code=400, detail="Số ngày cộng thêm phải lớn hơn 0.")

    db.execute(
        text(
            """
            INSERT INTO leave_year_adjustments (id, user_id, leave_year, extra_days, reason, created_by_id, created_at)
            VALUES (:id, :user_id, :leave_year, :extra_days, :reason, :created_by_id, :created_at)
            """
        ),
        {
            "id": str(uuid.uuid4()),
            "user_id": user_id,
            "leave_year": int(leave_year),
            "extra_days": int(extra_days),
            "reason": (reason or "").strip(),
            "created_by_id": user.id,
            "created_at": _dt_to_str(_now_vn()),
        },
    )
    db.commit()
    return RedirectResponse(url="/leave-schedule?msg=Đã cộng thêm ngày phép năm.", status_code=302)


@router.get("/export")
def export_leave_requests(
    request: Request,
    month: str = "",
    db: Session = Depends(get_db),
):
    _ensure_tables(db)
    user = login_required(request, db)
    if not (_can_adjust_after_cutoff(db, user) or _can_approve(db, user) or _can_admin_extra_days(db, user)):
        raise HTTPException(status_code=403, detail="Bạn không có quyền xuất dữ liệu đăng ký nghỉ.")

    target_month = month or _today_vn().strftime("%Y-%m")
    rows = db.execute(
        text(
            """
            SELECT *
            FROM leave_requests
            WHERE substr(start_date, 1, 7) = :target_month
            ORDER BY start_date ASC, group_label ASC, user_name ASC
            """
        ),
        {"target_month": target_month},
    ).fetchall()
    data = [_format_request_row(_row_to_dict(r)) for r in rows]

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Border, Font, Side
    except Exception as exc:
        raise HTTPException(
            status_code=500,
            detail="Thiếu thư viện openpyxl. Cần bổ sung openpyxl vào môi trường chạy ứng dụng để xuất Excel.",
        ) from exc

    wb = Workbook()
    ws = wb.active
    ws.title = "Dang ky nghi"

    ws["A1"] = "BẢNG TỔNG HỢP ĐĂNG KÝ NGHỈ"
    ws["A2"] = f"Tháng: {target_month}"
    ws.merge_cells("A1:K1")
    ws.merge_cells("A2:K2")
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"].font = Font(italic=True)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A2"].alignment = Alignment(horizontal="center")

    headers = [
        "STT", "Họ và tên", "Nhóm/Bộ phận", "Loại nghỉ", "Ký hiệu",
        "Từ ngày", "Đến ngày", "Số ngày công", "Lý do", "Trạng thái", "Người duyệt",
    ]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    row_no = 5
    for idx, item in enumerate(data, start=1):
        values = [
            idx,
            item.get("user_name"),
            item.get("group_label"),
            item.get("leave_type_label"),
            item.get("symbol"),
            item.get("start_date"),
            item.get("end_date"),
            item.get("day_count"),
            item.get("reason"),
            item.get("status_label"),
            item.get("approved_by_name") or "",
        ]
        for col_idx, value in enumerate(values, start=1):
            cell = ws.cell(row=row_no, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = border
        row_no += 1

    widths = {"A": 8, "B": 24, "C": 24, "D": 18, "E": 10, "F": 14, "G": 14, "H": 14, "I": 28, "J": 16, "K": 24}
    for col_name, width in widths.items():
        ws.column_dimensions[col_name].width = width

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"dang_ky_nghi_{target_month}.xlsx"
    headers = {"Content-Disposition": f'attachment; filename="{filename}"'}
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )
